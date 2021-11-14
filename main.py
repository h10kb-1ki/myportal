import streamlit as st
import requests
from bs4 import BeautifulSoup
import datetime
from PIL import Image
import io
import re
from pandas_datareader import data
import pandas as pd
from dateutil.relativedelta import relativedelta
import mplfinance as mpf
import altair as alt
import openpyxl as xl
import seaborn as sns
import matplotlib.pyplot as plt
import japanize_matplotlib

st.set_page_config(layout="wide")

'''
## MyPortal
'''

traffic = st.checkbox('Traffic')
if traffic:
    '''
    #### 東海道本線[豊橋～米原]
    '''
    url = 'https://transit.yahoo.co.jp/traininfo/detail/192/193/'
    res = requests.get(url)
    soup = BeautifulSoup(res.text, "html.parser")
    statusJ = soup.find('dd', class_='normal')
    if statusJ:
        st.write(statusJ.text)
    else:
        st.write('***遅延あり***')
    '''
    ###### ▶JR運行情報
    '''
    st.markdown('https://traininfo.jr-central.co.jp/zairaisen/status_detail.html?line=10001&lang=ja', unsafe_allow_html=True)
    st.write('')
    '''
    #### 名鉄名古屋本線
    '''
    url = 'https://transit.yahoo.co.jp/traininfo/detail/208/0/'
    res = requests.get(url)
    soup = BeautifulSoup(res.text, "html.parser")
    statusM = soup.find('dd', class_='normal')
    if statusM:
        st.write(statusM.text)
    else:
        st.write('***遅延あり***')
    ###### ▶名鉄（本線）運行情報
    '''
    st.markdown('https://top.meitetsu.co.jp/em/', unsafe_allow_html=True)
    #st.write('')

    '''
    ###### ▶名鉄バス（安城駅発 更生病院行）
    '''
    st.markdown('https://navi.meitetsu-bus.co.jp/mb/DepQR.aspx?p=320103000', unsafe_allow_html=True)

    '''
    ###### ▶乗り換え案内
    '''
    st.markdown('https://www.jorudan.co.jp/norikae/', unsafe_allow_html=True)

st.write('-----------------------------------------------------')
weather = st.checkbox('Weather')
if weather:
    nagoya = st.checkbox('名古屋市の天気')
    if nagoya == True:
        url = 'https://weathernews.jp/onebox/35.140631/136.856940/q=%E5%90%8D%E5%8F%A4%E5%B1%8B%E5%B8%82%E4%B8%AD%E5%B7%9D%E5%8C%BA&v=6e0846f392462de33f98b88b4ccdc67e48efadd56255e3c54a8f2bf8341c7f00&temp=c&lang=ja'
        res = requests.get(url)
        res.encoding = res.apparent_encoding
        soup = BeautifulSoup(res.text, "html.parser")

        #日付の取得（webページ上から）
        day_ = soup.find_all(class_='wTable__item')
        today = day_[6].text

        #最高・最低気温の取得
        kion = soup.find_all(class_='text wTable__item')
        max_today = kion[0].text
        min_today = kion[1].text
        max_tomorrow = kion[2].text
        min_tomorrow = kion[3].text

        #時間ごとの降水確率の取得
        kakuritu = soup.find_all(class_='text')
        today_p6 = kakuritu[2].text
        today_p12 = kakuritu[3].text
        today_p18 = kakuritu[4].text
        today_p24 = kakuritu[5].text
        tomorrow_p6 = kakuritu[8].text
        tomorrow_p12 = kakuritu[9].text
        tomorrow_p18 = kakuritu[10].text
        tomorrow_p24 = kakuritu[11].text

        #イメージアイコンの取得
        icon = soup.find_all(class_='day2Table__item weather')
        icon_today = 'https:' + icon[0].find('img').get('src')
        icon_tomorrow = 'https:' + icon[1].find('img').get('src')

        #解説コメントの取得
        title = soup.find(class_='tit-02').text
        info = soup.find(class_='comment no-ja')
        comment = info.text
        comment = comment.split('\n')[2]

        st.write(title)
        st.write(comment)
        st.write('■ '+ today)
        st.image(icon_today)
        st.write('最高気温:'+ max_today +'　最低気温:'+ min_today)
        st.write('～6時：'+ today_p6 +'　～12時：'+ today_p12 +'　～18時：'+ today_p18 +'　～24時：'+ today_p24)

        st.write('■ 明日の天気')
        st.image(icon_tomorrow)
        st.write('最高気温:'+ max_tomorrow +'　最低気温:'+ min_tomorrow)
        st.write('～6時：'+ tomorrow_p6 +'　～12時：'+ tomorrow_p12 +'　～18時：'+ tomorrow_p18 +'　～24時：'+ tomorrow_p24)

    anjo = st.checkbox('安城市の天気')
    if anjo == True:
        url = 'https://weathernews.jp/onebox/34.948663/137.079025/q=%E6%84%9B%E7%9F%A5%E7%9C%8C%E5%AE%89%E5%9F%8E%E5%B8%82&v=3fa1edac9382759435af39576ac457ebaf29245456fafb5ff44b458182f4cbbc&temp=c&lang=ja'
        res = requests.get(url)
        res.encoding = res.apparent_encoding
        soup = BeautifulSoup(res.text, "html.parser")

        #日付の取得（webページ上から）
        day_ = soup.find_all(class_='wTable__item')
        today = day_[6].text

        #最高・最低気温の取得
        kion = soup.find_all(class_='text wTable__item')
        max_today = kion[0].text
        min_today = kion[1].text
        max_tomorrow = kion[2].text
        min_tomorrow = kion[3].text

        #時間ごとの降水確率の取得
        kakuritu = soup.find_all(class_='text')
        today_p6 = kakuritu[2].text
        today_p12 = kakuritu[3].text
        today_p18 = kakuritu[4].text
        today_p24 = kakuritu[5].text
        tomorrow_p6 = kakuritu[8].text
        tomorrow_p12 = kakuritu[9].text
        tomorrow_p18 = kakuritu[10].text
        tomorrow_p24 = kakuritu[11].text

        #イメージアイコンの取得
        icon = soup.find_all(class_='day2Table__item weather')
        icon_today = 'https:' + icon[0].find('img').get('src')
        icon_tomorrow = 'https:' + icon[1].find('img').get('src')

        #解説コメントの取得
        title = soup.find(class_='tit-02').text
        info = soup.find(class_='comment no-ja')
        comment = info.text
        comment = comment.split('\n')[2]

        st.write(title)
        st.write(comment)
        st.write('■ '+ today)
        st.image(icon_today)
        st.write('最高気温:'+ max_today +'　最低気温:'+ min_today)
        st.write('～6時：'+ today_p6 +'　～12時：'+ today_p12 +'　～18時：'+ today_p18 +'　～24時：'+ today_p24)

        st.write('■ 明日の天気')
        st.image(icon_tomorrow)
        st.write('最高気温:'+ max_tomorrow +'　最低気温:'+ min_tomorrow)
        st.write('～6時：'+ tomorrow_p6 +'　～12時：'+ tomorrow_p12 +'　～18時：'+ tomorrow_p18 +'　～24時：'+ tomorrow_p24)

    '''
    ###### ▶雨雲レーダー
    '''
    st.markdown('https://tenki.jp/radar/map/', unsafe_allow_html=True)

st.write('-----------------------------------------------------')
news = st.checkbox('NEWS')
if news:
    yahoo = st.checkbox('Yahoo! ニュース トピックス')
    if yahoo == True:
        url = 'https://www.yahoo.co.jp/'
        r = requests.get(url)
        soup = BeautifulSoup(r.text, 'html.parser')
        elems = soup.find_all(href = re.compile('news.yahoo.co.jp/pickup'))
        for i in range(0, len(elems)):
            # titleを取得
            title = elems[i].text
            st.write(f'・{title}')
            # linkを取得
            link = elems[i].attrs['href']
            st.write(f'>{link}')
            #st.write('')
    seiyaku = st.checkbox('製薬業界ニュース')
    if seiyaku ==True:
        url = 'https://answers.ten-navi.com/pharmanews/pharma_category/1/'
        res = requests.get(url)
        res.encoding = res.apparent_encoding
        soup = BeautifulSoup(res.text, "html.parser")

        titles = soup.find_all('h2')
        tag = soup.find_all(class_='tag')
        ref = soup.find_all('a', class_='clearfix')

        for i in range(0, len(titles)):
            if tag[i].text == 'ニュース解説':
                title = titles[i].text
                link = ref[i].attrs['href']
                st.write(f'・{title}')
                st.write(f'>{link}')
                #st.write('')
                
st.write('-----------------------------------------------------')                
Finance = st.checkbox('Finance')
if Finance:
    st.set_option('deprecation.showPyplotGlobalUse', False)
    dic = {'サントリー': '2587.JP',
      'アサヒ': '2502.JP',
      'キリン': '2503.JP',
      'サッポロ': '2501.JP',
      'タケダ': '4502.JP',
      'アステラス': '4503.JP',
      '大塚': '4578.JP',
      '第一三共': '4568.JP',
      'エーザイ': '4523.JP',
      '中外': '4519.JP',
      '大日本住友': '4506.JP',
      '塩野義': '4507.JP',
      '協和キリン': '4151.JP',
      '小野薬品': '4528.JP'}
    name = list(dic.keys())
    today = datetime.datetime.now()
    start_point = st.selectbox('開始', ('1ヶ月前', '3ヶ月前', '半年前', '1年前', '任意'), index=2)
    if start_point == '1ヶ月前':
        start = today - relativedelta(months=1)
    elif start_point == '3ヶ月前':
        start = today - relativedelta(months=3)
    elif start_point == '半年前':
        start = today - relativedelta(months=6)
    elif start_point == '1年前':
        start = today - relativedelta(months=12) 
    else:
        start = st.date_input('開始')

    end = st.date_input('終了')
    company_name = st.selectbox('銘柄', name, index=11)
    company_code = dic[company_name]

    df = data.DataReader(company_code, 'stooq', start, end)
    df = df.sort_values('Date', ascending=True)

    cs  = mpf.make_mpf_style(gridcolor="lightgray", facecolor="white", edgecolor="#202426", figcolor="white", 
            rc={"xtick.color":"black", "xtick.labelsize":12, 
                "ytick.color":"black", "ytick.labelsize":12, 
                "axes.labelsize":15, "axes.labelcolor":"black"})
    fig = mpf.plot(df, type='candle', volume=True, mav=(5, 25, 50), figratio=(12,4), style=cs)
    st.pyplot(fig)

st.write('-----------------------------------------------------')  
hobby = st.checkbox('Hobby & Health')
if hobby:
    '''
    ###### ▶colt python
    '''
    st.write('準備中...')

    '''
    ###### ▶世界は もっと 面白くていい！
    '''
    st.markdown('https://twitter.com/k5dbzrmjne77i5r', unsafe_allow_html=True)
    
    '''
    ###### ▶Pep Up
    '''
    st.markdown('https://pepup.life/home', unsafe_allow_html=True)

st.write('-----------------------------------------------------')                
MyLib = st.checkbox('Library')
if MyLib:
    '''
    ###### ▶VCM血中濃度シミュレーション
    '''
    st.markdown('https://share.streamlit.io/h10kb-1ki/vcm/vcm_para.py', unsafe_allow_html=True)
    '''
    ###### ▶業務分析アプリ
    '''
    st.markdown('https://share.streamlit.io/h10kb-1ki/workanalysis/WorkAnalysis.py', unsafe_allow_html=True)
    '''
    ###### ▶ToDoリスト（閉鎖中）
    '''
    st.markdown('https://share.streamlit.io/h10kb-1ki/todolist/ToDoList.py', unsafe_allow_html=True)

    '''
    ###### ▶一郎の部屋
    '''
    st.markdown('https://sites.google.com/view/tdmichiro/%E3%83%9B%E3%83%BC%E3%83%A0', unsafe_allow_html=True)
    
    kampo = st.checkbox('漢方比較')
    if kampo == True:
        wb = xl.load_workbook('kampo.xlsx')
        ws = []
        for i in wb.worksheets:
            ws.append(i.title)

        df = pd.read_excel('kampo.xlsx', sheet_name=ws[0], header=0, index_col=0)
        for j in range(1, len(ws)):
            df_sheet = pd.read_excel('kampo.xlsx', sheet_name=ws[j], header=0, index_col=0)
            df = df.merge(df_sheet, how='outer', left_index=True, right_index=True)

        st.title('漢方：含有生薬の比較')
        st.write('1日量（ツムラ製品、通常量）中の生薬含有量を表示')
        kampo_list = sorted(df.columns)
        selection = st.multiselect('漢方を選択', kampo_list)

        df = df[selection]
        #空白行(Nan)を削除
        df.dropna(subset=selection, how='all', inplace=True)
        fig, ax = plt.subplots(1, 1, figsize=(12, 10))
        ax = sns.heatmap(df, annot=True, fmt='.1f', cmap='Blues', vmax=10, vmin=0, ax=ax)

        btn = st.button('表示')

        if btn == True:
            st.pyplot(fig)
